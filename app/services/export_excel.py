"""Excel export for the Download screen — openpyxl, a Summary sheet plus a flat
BOQ Detail sheet (room x material rows)."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.boq import BOQResponse

HEADER_FILL = PatternFill(start_color="36355B", end_color="36355B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CURRENCY_FORMAT = "#,##0.00"


def _style_header_row(ws, row_number: int) -> None:
    for cell in ws[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 40)


def generate_boq_excel(boq: BOQResponse) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Project", boq.project_name])
    summary.append(["Location", boq.location])
    summary.append(["Generated At", boq.generated_at.isoformat()])
    summary.append([])
    summary.append(["Room", "Room Total Cost"])
    _style_header_row(summary, summary.max_row)

    for room in boq.rooms:
        summary.append([room.room_name, room.room_total_cost])
        summary.cell(row=summary.max_row, column=2).number_format = CURRENCY_FORMAT

    summary.append([])
    summary.append(["Total Cost", boq.total_cost])
    summary.cell(row=summary.max_row, column=1).font = Font(bold=True)
    summary.cell(row=summary.max_row, column=2).font = Font(bold=True)
    summary.cell(row=summary.max_row, column=2).number_format = CURRENCY_FORMAT
    _autosize(summary)

    detail = wb.create_sheet("BOQ Detail")
    detail.append(
        [
            "Room",
            "Room Type",
            "Material",
            "Theoretical Qty",
            "Correction Factor",
            "Confidence",
            "Quantity",
            "Unit",
            "Rate/Unit",
            "Total Cost",
        ]
    )
    _style_header_row(detail, 1)

    for room in boq.rooms:
        for material in room.materials:
            detail.append(
                [
                    room.room_name,
                    room.room_type.value,
                    material.material_name,
                    material.theoretical_quantity,
                    material.correction_factor,
                    material.correction_confidence.value,
                    material.quantity,
                    material.unit,
                    material.rate_per_unit,
                    material.total_cost,
                ]
            )
            row = detail.max_row
            detail.cell(row=row, column=9).number_format = CURRENCY_FORMAT
            detail.cell(row=row, column=10).number_format = CURRENCY_FORMAT
    _autosize(detail)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
